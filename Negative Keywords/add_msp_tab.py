"""
Extract keywords from 6 general/MSP tabs, deduplicate, and add as
'MSP Launch All Negative Keywords List' tab to compiled_negkw_by_region.xlsx.
"""
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
import re

XLSX_IN   = "MSPL_negkw_full.xlsx"
XLSX_OUT  = "compiled_negkw_by_region.xlsx"
NEW_TAB   = "MSP Launch - All Neg Keywords"

# Cells/values to skip
SKIP_EXACT = {
    "broad", "phrase", "exact", "negative keywords",
    "general neg keyword list", "lead gen search terms",
    "local it services providers negative keywords list",
    "pay per click master negative keywords list",
    "job related", "research & stats", "education/groups", "deals",
    "price, quote & diy", "manufacturers & industrial", "materials", "legal",
    "create the list for the uk, then ca and then aus using:",
    "local seo with 5 keywords", "global seo with 5 keywords",
    "research on google", "elements from the us list",
    "outscraper", "louis@msplaunchpad.com", "slangen97",
}
SKIP_STARTS = ["http", "www.", "create ", "local seo", "global seo",
               "research on", "elements from"]

URL_RE = re.compile(r"https?://|www\.", re.I)

def should_skip(text):
    t = str(text).strip()
    if not t or len(t) <= 1:
        return True
    tl = t.lower()
    if tl in SKIP_EXACT:
        return True
    if URL_RE.search(t):
        return True
    for s in SKIP_STARTS:
        if tl.startswith(s):
            return True
    # Skip pure numbers / addresses / timestamps
    if re.match(r"^[\d\s\-\+\(\)\.\:\/]+$", t):
        return True
    return False

def clean(text):
    t = str(text).strip()
    t = t.replace("“", '"').replace("”", '"')
    t = t.replace("‘", "'").replace("’", "'")
    t = re.sub(r"\s+", " ", t)
    return t.strip()

# ── Tab-specific extractors ─────────────────────────────────────────────────

def extract_single_column(ws):
    """For General_Keywords_3.0, General Neg keys, MSP Launchpad Proprietary Negat."""
    kws = set()
    for row in ws.iter_rows(values_only=True):
        cell = row[0] if row else None
        if cell is None:
            continue
        val = clean(str(cell))
        # Some entries have double-wrapped quotes: ""keyword"" → "keyword"
        val = re.sub(r'^""+(.+)""+$', r'"\1"', val)
        val = re.sub(r"^'+(.+)'+$", r'"\1"', val)
        if val and not should_skip(val):
            kws.add(val.lower())
    return kws

def extract_negative_keywords_tab(ws):
    """
    Negative-Keywords tab: 3 column groups, each with keyword + BROAD/PHRASE label.
    Keyword cols: 0, 3, 6. Label cols: 1, 4, 7.
    """
    kws = set()
    KW_COLS = [0, 3, 6]
    LABEL_COLS = {1: "broad", 4: "phrase", 7: "broad"}

    for row in ws.iter_rows(values_only=True):
        row = list(row)
        for kw_col in KW_COLS:
            if kw_col >= len(row) or row[kw_col] is None:
                continue
            kw = clean(str(row[kw_col]))
            if not kw or should_skip(kw):
                continue
            # Get match type from adjacent label column
            label_col = kw_col + 1
            label = ""
            if label_col < len(row) and row[label_col]:
                label = str(row[label_col]).strip().lower()
            # Format keyword by match type
            if label == "phrase" and not (kw.startswith('"') and kw.endswith('"')):
                kw = f'"{kw}"'
            elif label == "exact" and not (kw.startswith('[') and kw.endswith(']')):
                kw = f'[{kw}]'
            kws.add(kw.lower())
    return kws

def extract_major_msps(ws):
    """Major MSPs List - USA: company names in column A only."""
    kws = set()
    for row in ws.iter_rows(values_only=True):
        cell = row[0] if row else None
        if cell is None:
            continue
        val = clean(str(cell))
        if val and not should_skip(val):
            kws.add(val.lower())
    return kws

def extract_general_neg_keys_2(ws):
    """
    General Neg Keys 2.0: multi-column, all phrase match (in quotes).
    Skip the category header row.
    """
    kws = set()
    for row in ws.iter_rows(values_only=True):
        for cell in row:
            if cell is None:
                continue
            val = clean(str(cell))
            if not val or should_skip(val):
                continue
            # Already in quotes mostly; normalise
            val = re.sub(r'^""+(.+)""+$', r'"\1"', val)
            kws.add(val.lower())
    return kws

# ── Main ────────────────────────────────────────────────────────────────────

def main():
    wb_src = openpyxl.load_workbook(XLSX_IN, read_only=True, data_only=True)

    extractors = {
        "General_Keywords_3.0":          (extract_single_column,       "General Keywords 3.0"),
        "General Neg keys":              (extract_single_column,       "General Neg Keys"),
        "Negative-Keywords":             (extract_negative_keywords_tab,"Negative Keywords"),
        "MSP Launchpad Proprietary Negat":(extract_single_column,      "MSP Launchpad Proprietary"),
        "Major MSPs List - USA":         (extract_major_msps,          "Major MSPs List USA"),
        "General Neg Keys 2.0":          (extract_general_neg_keys_2,  "General Neg Keys 2.0"),
    }

    all_keywords = set()
    source_data  = []   # [(source_label, kw), ...]

    print(f"\nExtracting from 6 MSP/general tabs...\n")

    for tab_name, (extractor, label) in extractors.items():
        ws = wb_src[tab_name]
        kws = extractor(ws)
        new_kws = kws - all_keywords
        all_keywords.update(new_kws)
        for kw in sorted(new_kws, key=lambda x: x.lstrip('"[').lower()):
            source_data.append((label, kw))
        print(f"  {tab_name:<40} | {len(kws):>5} tab-unique | {len(new_kws):>5} new after dedup")

    # ── Load existing output workbook and add new tab ──
    wb_out = openpyxl.load_workbook(XLSX_OUT)

    # Remove existing tab if re-running
    if NEW_TAB in wb_out.sheetnames:
        del wb_out[NEW_TAB]

    ws_new = wb_out.create_sheet(title=NEW_TAB, index=0)  # put it first

    # Header styling — dark teal
    TEAL = "1F6B75"
    hdr_fill = PatternFill("solid", fgColor=TEAL)
    hdr_font = Font(bold=True, color="FFFFFF", size=11)

    headers = ["Source", "Negative Keyword"]
    for col, h in enumerate(headers, 1):
        cell = ws_new.cell(row=1, column=col, value=h)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="left")

    ws_new.column_dimensions["A"].width = 30
    ws_new.column_dimensions["B"].width = 60
    ws_new.freeze_panes = "A2"

    for r, (source, kw) in enumerate(source_data, 2):
        ws_new.cell(row=r, column=1, value=source)
        ws_new.cell(row=r, column=2, value=kw)

    wb_out.save(XLSX_OUT)

    print(f"\n{'='*60}")
    print(f"  Tab added: '{NEW_TAB}'")
    print(f"  Total unique keywords: {len(all_keywords):,}")
    print(f"  Saved to: {XLSX_OUT}")
    print(f"{'='*60}\n")

if __name__ == "__main__":
    main()
