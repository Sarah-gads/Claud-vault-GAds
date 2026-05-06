"""
Compile negative keywords grouped by geographic region into separate Excel tabs.
Deduplicates within each region tab.
"""
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from collections import defaultdict
import re

XLSX_IN  = "MSPL_negkw_full.xlsx"
XLSX_OUT = "compiled_negkw_by_region.xlsx"

SKIP_TABS = {
    "General_Keywords_3.0", "Notes", "General Neg keys", "Negative-Keywords",
    "AdCopies", "MSP Launchpad Proprietary Negat", "Major MSPs List - USA",
    "SOP - Hospitals, Universities, ", "General Neg Keys 2.0", "Sheet10",
}

# Map each client tab → region
REGION_MAP = {
    "365 Managed IT":       "US",
    "AJTC":                 "US",
    "AlwaysOnIT":           "US",
    "Carmichael":           "US",
    "Capstone Works":       "US",
    "DivergeIT":            "US",
    "Infotank":             "US",
    "LEET Services":        "US",
    "Red Team IT":          "US",
    "SageNetCom":           "US",
    "Techlocity":           "US",
    "Trinity Networx":      "US",
    "Version2 LLC":         "US",
    "MBPS":                 "US",
    "Web IT":               "US",
    "Sterling Tech":        "US",
    "Citadel Blue":         "US",
    "UDNI":                 "US",
    "Vital Integrators":    "US",
    "Intelligent IT":       "US",
    "ArgoCTS":              "US",
    "Nuage ":               "US",
    "CyberFire IT":         "US",
    "Unified Technicians":  "US",

    "Serveline":            "UK",
    "Sonar IT":             "UK",
    "Netflo":               "UK",
    "AccrueTek":            "UK",
    "Anantek":              "UK",

    "EPIC IT":              "Australia",
    "Soma Tech":            "Australia",
    "epochLABS":            "Australia",

    "RTC":                  "Canada",
    "Infoware":             "Canada",
    "CloudOrbis":           "Canada",

    "OxygenIT":             "New Zealand",
    "Cover Me":             "France",
}

# Region display order and tab colors (hex fill)
REGION_ORDER = ["US", "UK", "Australia", "Canada", "New Zealand", "France"]
REGION_COLORS = {
    "US":          "4472C4",   # blue
    "UK":          "C00000",   # red
    "Australia":   "70AD47",   # green
    "Canada":      "ED7D31",   # orange
    "New Zealand": "7030A0",   # purple
    "France":      "00B0F0",   # light blue
}

URL_RE = re.compile(r"https?://|www\.", re.I)
HEADER_FRAGMENTS = [
    "local institutions", "local it services", "competitors", "excluded negative",
    "general neg", "excluded search", "school / university", "phrase match",
    "keyword", "website", "company", "competitors list", "competitors - msp",
    "excluded competitors", "general keywords", "mspl |", "mspl|",
    "free match type", "negative keys folder", "negative keyword list",
    "pay per click", "job related", "research", "education", "deals",
    "broad", "phrase", "exact", "location:", "general neg keys",
    "genaral negative", "general negative keywords",
]

def is_skip(text):
    t = str(text).strip()
    if not t or len(t) <= 1:
        return True
    if URL_RE.search(t):
        return True
    tl = t.lower()
    if any(f in tl for f in HEADER_FRAGMENTS):
        return True
    if tl.startswith("location"):
        return True
    if re.match(r"^[\d\s\-\+\(\)\.]+$", t):
        return True
    return False

def clean(text):
    text = str(text).strip()
    text = text.replace("“", '"').replace("”", '"')
    text = text.replace("‘", "'").replace("’", "'")
    text = re.sub(r"\s+", " ", text)
    return text.strip()

def extract_location(ws):
    first_row = next(ws.iter_rows(values_only=True, max_row=1), [])
    for cell in first_row:
        if cell and str(cell).lower().strip().startswith("location"):
            loc = re.sub(r"location\s*:\s*", "", str(cell), flags=re.I).strip()
            return loc
    return "Unknown"

def extract_keywords(ws):
    keywords = set()
    for row in ws.iter_rows(values_only=True):
        for cell in row:
            if cell is None:
                continue
            raw = clean(str(cell))
            for part in raw.split("\n"):
                part = clean(part)
                if part and not is_skip(part):
                    keywords.add(part.lower())
    return keywords

def write_region_sheet(wb, region, rows, color_hex):
    ws = wb.create_sheet(title=region)

    # Header styling
    hdr_fill = PatternFill("solid", fgColor=color_hex)
    hdr_font = Font(bold=True, color="FFFFFF", size=11)

    headers = ["Client", "Location", "Negative Keyword"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="left")

    # Column widths
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 55

    for r, (client, loc, kw) in enumerate(rows, 2):
        ws.cell(row=r, column=1, value=client)
        ws.cell(row=r, column=2, value=loc)
        ws.cell(row=r, column=3, value=kw)

    # Freeze header row
    ws.freeze_panes = "A2"
    return len(rows)

def main():
    wb_in = openpyxl.load_workbook(XLSX_IN, read_only=True, data_only=True)
    wb_out = openpyxl.Workbook()
    wb_out.remove(wb_out.active)  # remove default sheet

    # Collect keywords per region (deduplicated across clients in same region)
    region_data = defaultdict(list)  # region → [(client, location, kw), ...]
    region_kw_seen = defaultdict(set)  # region → set of kws already added

    client_tabs = [n for n in wb_in.sheetnames if n not in SKIP_TABS]

    print(f"\nExtracting from {len(client_tabs)} client tabs...\n")

    for tab in client_tabs:
        region = REGION_MAP.get(tab, "Other")
        ws = wb_in[tab]
        location = extract_location(ws)
        keywords = extract_keywords(ws)

        added = 0
        for kw in sorted(keywords, key=lambda x: x.lstrip('"[').lower()):
            if kw not in region_kw_seen[region]:
                region_kw_seen[region].add(kw)
                region_data[region].append((tab, location, kw))
                added += 1

        print(f"  {tab:<30} | {region:<12} | {len(keywords):>5} tab-unique | {added:>5} added to region")

    # Write one sheet per region
    print(f"\nWriting Excel file...\n")
    total = 0
    for region in REGION_ORDER:
        if region not in region_data:
            continue
        rows = region_data[region]
        color = REGION_COLORS.get(region, "808080")
        count = write_region_sheet(wb_out, region, rows, color)
        total += count
        print(f"  {region:<15} | {count:>6,} unique keywords")

    # Handle any unmapped regions
    for region, rows in region_data.items():
        if region not in REGION_ORDER:
            color = "808080"
            count = write_region_sheet(wb_out, region, rows, color)
            total += count
            print(f"  {region:<15} | {count:>6,} unique keywords")

    wb_out.save(XLSX_OUT)

    print(f"\n{'='*55}")
    print(f"  Total keywords written: {total:,}")
    print(f"  Output: {XLSX_OUT}")
    print(f"{'='*55}\n")

if __name__ == "__main__":
    main()
