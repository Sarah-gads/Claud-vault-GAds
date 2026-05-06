"""
Compile negative keywords from all client tabs in MSPL_negkw_full.xlsx.
Handles all 4 column formats found across tabs.
Deduplicates within each tab. Outputs one master CSV.
"""
import openpyxl
import csv
import re
from collections import defaultdict

XLSX = "MSPL_negkw_full.xlsx"
OUTPUT = "compiled_all_clients_negkw.csv"

# Reference/utility tabs to skip — not client-specific
SKIP_TABS = {
    "General_Keywords_3.0", "Notes", "General Neg keys", "Negative-Keywords",
    "AdCopies", "MSP Launchpad Proprietary Negat", "Major MSPs List - USA",
    "SOP - Hospitals, Universities, ", "General Neg Keys 2.0", "Sheet10",
}

# These look like URLs
URL_RE = re.compile(r"https?://|www\.", re.I)

# Strings that are structural headers, not keywords
HEADER_FRAGMENTS = [
    "local institutions", "local it services", "competitors", "excluded negative",
    "general neg", "excluded search", "school / university", "phrase match",
    "keyword", "website", "company", "competitors list", "competitors - msp",
    "excluded competitors", "general keywords", "mspl |", "mspl|",
    "free match type", "negative keys folder", "negative keyword list",
    "pay per click", "job related", "research", "education", "deals",
    "broad", "phrase", "exact", "location:", "general neg keys",
    "genaral negative", "general negative keywords",
]

def is_header(text):
    t = text.lower().strip()
    for frag in HEADER_FRAGMENTS:
        if frag in t:
            return True
    return False

def is_url(text):
    return bool(URL_RE.search(text.strip()))

def is_location_line(text):
    return text.strip().lower().startswith("location")

def clean(text):
    if not text:
        return ""
    text = str(text).strip()
    # Normalize smart quotes and encoding artifacts
    text = text.replace("“", '"').replace("”", '"')
    text = text.replace("‘", "'").replace("’", "'")
    text = re.sub(r"\s+", " ", text)
    return text.strip()

def should_skip(text):
    t = text.strip()
    if not t or len(t) <= 1:
        return True
    if is_url(t):
        return True
    if is_header(t):
        return True
    if is_location_line(t):
        return True
    # Skip obvious non-keywords: pure numbers, addresses, phone numbers
    if re.match(r"^[\d\s\-\+\(\)\.]+$", t):
        return True
    return False

def extract_location(ws):
    """Pull the LOCATION line from row 0."""
    first_row = next(ws.iter_rows(values_only=True, max_row=1), [])
    for cell in first_row:
        if cell and str(cell).lower().strip().startswith("location"):
            loc = str(cell).strip()
            # Clean up "LOCATION: " prefix
            loc = re.sub(r"location\s*:\s*", "", loc, flags=re.I).strip()
            return loc
    return "Unknown"

def extract_keywords_from_tab(ws):
    """
    Extract all negative keywords from a sheet regardless of column format.
    Returns a set of cleaned, deduplicated keyword strings.
    """
    keywords = set()
    for row in ws.iter_rows(values_only=True):
        for cell in row:
            if cell is None:
                continue
            # A cell may have multiple keywords separated by newlines
            raw = clean(str(cell))
            for part in raw.split("\n"):
                part = clean(part)
                if not part or should_skip(part):
                    continue
                # Normalize case: keep exact/phrase match formatting,
                # lowercase everything else
                if part.startswith('"') or part.startswith('['):
                    keywords.add(part.lower())
                else:
                    keywords.add(part.lower())
    return keywords

def main():
    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)

    results = []  # list of (tab_name, location, keyword)

    client_tabs = [name for name in wb.sheetnames if name not in SKIP_TABS]

    print(f"\nProcessing {len(client_tabs)} client tabs...\n")

    for tab_name in client_tabs:
        ws = wb[tab_name]
        location = extract_location(ws)
        keywords = extract_keywords_from_tab(ws)

        print(f"  {tab_name:<30} | {location:<35} | {len(keywords):>5} unique keywords")

        for kw in sorted(keywords, key=lambda x: x.lstrip('"[').lower()):
            results.append((tab_name, location, kw))

    # Write master CSV
    with open(OUTPUT, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["Client / Tab", "Location", "Negative Keyword"])
        current_tab = None
        for tab, loc, kw in results:
            if tab != current_tab:
                writer.writerow([])  # blank separator between clients
                writer.writerow([f"=== {tab} — {loc} ===", "", ""])
                current_tab = tab
            writer.writerow([tab, loc, kw])

    total = len(results)
    print(f"\n{'='*70}")
    print(f"  TOTAL keywords across all {len(client_tabs)} client tabs: {total:,}")
    print(f"  Output: {OUTPUT}")
    print(f"{'='*70}\n")

if __name__ == "__main__":
    main()
